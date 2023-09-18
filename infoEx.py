import firstScene
import showInfo
import test
import sys
import os
import webbrowser
import successWindow
import openpyxl
from PyQt5 import QtCore, QtGui, QtWidgets
from PyQt5.QtWidgets import QApplication

def readFile(dire):
	f = open(dire,'r',encoding = 'utf-8')
	txt = f.read()
	f.close()
	return txt

def OpenFile():
	try:
		filepath,f = QtWidgets.QFileDialog.getOpenFileName(scene1,'choose file','','text files(*.txt)')
		s1ui.filepathTxt.setText(filepath)
		txt = readFile(filepath)
		s1ui.displaytxt.setText(txt)
		# the path of choosed file
		f = open('./temp/dir.txt','w')
		f.write(filepath)
		f.close()
	except:
		s1ui.displaytxt.setText("no file choosed")

def ChoosePath():
	try:
		filepath = QtWidgets.QFileDialog.getExistingDirectory(scene2,'choose file')
		ui2.textBrowser.setText(filepath)
		f = open('./temp/choosePathInfo.txt','w',encoding = 'utf-8')
		f.write(filepath)
		f.close()
	except:
		ui2.textBrowser.setText("no file choosed")

def ShowNER():
	try:
		f = webbrowser.open_new_tab("file:///"+os.getcwd()+"/temp/h.html")
	except:
		print('error showNER')

def SaveNER():
	try:
		f = open('./temp/choosePathInfo.txt','r',encoding = 'utf-8')
		path = f.read()+"/NER.html"
		f.close()
		f = open("./temp/h.html",'r',encoding = 'utf-8')
		txt = f.read()
		f.close()
		f1 = open(path,'w',encoding='utf-8')
		f1.write(txt)
		f1.close()
		ui2.textBrowser.setText('successfully save to '+path)
	except:
		ui2.textBrowser.setText('please choose path!!!')


def ShowInfo():
	s1ui.displaytxt.setText('runing model... please wait..')
	f = open('./temp/dir.txt','r',encoding = 'utf-8')
	path = f.read()
	f.close()
	f = open(path,'r',encoding = 'utf-8')
	txt = f.read()
	df = test.runModelHtml(txt)
	ui2.showNerButton.clicked.connect(ShowNER)
	ui2.saveNerButton.clicked.connect(SaveNER)
	ui2.choosePathButton.clicked.connect(ChoosePath)
	showExcel()
	scene2.show()
	# try:
	# 	# pic = QtGui.QPixmap("./temp/test.jpg")
	# 	# ui2.showImage.setPixmap(pic)
	# 	# ui2.showImage.setScaledContents(True)
	# 	f = open('./temp/choosePathInfo.txt','r',encoding = 'utf-8')
	# 	path = f.read()
	# 	f.close()
	# 	f = open(path,'r',encoding = 'utf-8')
	# 	txt = f.read()
	# 	df = test.runModelHtml(txt)

	# 	ui2.showNerButton.clicked.connect(ShowNER)
	# 	ui2.saveNerButton.clicked.connect(SaveNER)
	# 	ui2.choosePathButton.clicked.connect(ChoosePath)
	# 	showExcel()
	# 	scene2.show()
	# except:
	# 	print("error ShowInfo")

def showExcel():
	wb = openpyxl.load_workbook("./temp/info.xlsx") 
	ws = wb.worksheets[0]
	rows = ws.max_row
	columns = ws.max_column
	ui2.showExcel.setColumnCount(columns)
	ui2.showExcel.setRowCount(rows)
	for row in range(rows):
		for column in range(columns):
			cell = ws.cell(row=row+1, column=column+1)
			ui2.showExcel.setItem(row, column, QtWidgets.QTableWidgetItem(str(cell.value)))
def excel():
	s1ui.displaytxt.setText('saving excel...')
	f = open('./temp/dir.txt','r',encoding = 'utf-8')
	path = f.read()
	f.close()
	f = open(path,'r',encoding = 'utf-8')
	txt = f.read()
	df = test.runModel(txt)
	df.to_excel('./infoResult.xlsx')
	s1ui.displaytxt.setText('successfully saving to ./infoResult.xlsx')
	# succUI.showPath.setText('./infoResult.xlsx')
	# succScene.show()

if __name__ == '__main__':
	app = QApplication(sys.argv)
	scene1 = QtWidgets.QMainWindow()
	s1ui = firstScene.Ui_MainWindow()
	s1ui.setupUi(scene1)
	s1ui.chooseFileButton.clicked.connect(OpenFile)
	s1ui.exExcelButton.clicked.connect(excel)
	s1ui.showInfoButton.clicked.connect(ShowInfo)
	# show info window
	scene2 = QtWidgets.QMainWindow()
	ui2 = showInfo.Ui_MainWindow()
	ui2.setupUi(scene2)
	scene1.show()
	sys.exit(app.exec_())
	# success window
	succScene = QtWidgets.QMainWindow()
	succUI = successWindow.Ui_MainWindow()
	succUI.setupUi(scene1)
	#
	errWindow = QtWidgets.QWidget()
	errWindow.resize(400,200)
	errWindow.setWindowTitle("Fail to Save")
	errWindow.label = QtWidgets.QLabel(errWindow)
	errWindow.label.setText("Fail to save, \ncheck if choose the path to save")